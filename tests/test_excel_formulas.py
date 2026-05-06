from datetime import date
from decimal import Decimal

from openpyxl import load_workbook

from models import Order, OrderItem
from services.excel_exporter import ExcelExporter


def test_exporter_writes_expected_formulas(tmp_path) -> None:
    order = Order(
        order_number="123",
        order_date=date(2026, 5, 5),
        sender="Test",
        dispatch_city="Moscow",
        tenge_rate=Decimal("6"),
        tenge_rate_fact=Decimal("6.1"),
        delivery_percent=Decimal("6"),
        expenses=Decimal("10"),
        items=[
            OrderItem(
                source_number="1",
                full_name="Иван Иванов",
                normalized_name="иван иванов",
                group_id=1,
                group_name="гр. Тест",
                amount_tenge=Decimal("12000"),
                discount_tenge=Decimal("1000"),
                amount_with_discount_tenge=Decimal("11000"),
                registration_fee=Decimal("500"),
                delivery_percent=Decimal("6"),
                transferred_rub=Decimal("7000"),
            )
        ],
    )
    path = tmp_path / "out.xlsx"
    ExcelExporter().export_new(order, path)
    ws = load_workbook(path, data_only=False).active
    assert ws["F4"].value == "=C4-E4"
    assert ws["G4"].value == "=F4/$R$2"
    assert ws["H4"].value == "=(C4-D4)/$R$2*$H$2"
    assert ws["D4"].value == 500
    assert ws["P4"].value is None
    assert ws["M4"].value == 12000
    assert ws["D5"].value == "=SUM(D4:D4)"
    assert ws["P5"].value == "=SUM(Y4:Y4)"
    assert ws["K2"].value == "=(G7+T2)/G7-1"
    assert ws["K4"].value == "=G4*$K$2"
    assert ws["T2"].value == "=SUM(T3:T6)"
    assert ws["C5"].font.color.rgb == "FF2E75B5"
    assert ws["E5"].font.color.rgb == "FF00B050"
    assert ws["B8"].value == "Номер заказа:"
    assert ws["B12"].value == "Оплатил:"
    assert ws["E15"].value == 0
    assert ws["F12"].font.color.rgb == "FF2E75B5"
    assert ws["H12"].font.color.rgb == "FFC00000"
    assert ws["I12"].font.color.rgb == "FF00B050"
    assert ws["B4"].alignment.horizontal == "left"
    assert ws["B1"].alignment.horizontal == "center"
    assert ws["F12"].font.bold is True
