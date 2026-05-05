from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook


@dataclass(slots=True)
class TemplateCellInfo:
    coordinate: str
    value: str | int | float | None
    formula: str | None
    style_id: int


@dataclass(slots=True)
class TemplateAnalysis:
    path: Path
    sheet_title: str
    max_row: int
    max_column: int
    freeze_panes: str | None
    merged_ranges: list[str] = field(default_factory=list)
    column_widths: dict[str, float] = field(default_factory=dict)
    formulas: dict[str, str] = field(default_factory=dict)
    non_empty_cells: list[TemplateCellInfo] = field(default_factory=list)


class ExcelTemplateReader:
    def analyze(self, path: Path) -> TemplateAnalysis:
        wb = load_workbook(path)
        ws = wb[wb.sheetnames[0]]
        analysis = TemplateAnalysis(
            path=path,
            sheet_title=ws.title,
            max_row=ws.max_row,
            max_column=ws.max_column,
            freeze_panes=str(ws.freeze_panes) if ws.freeze_panes else None,
            merged_ranges=[str(rng) for rng in ws.merged_cells.ranges],
        )
        for key, dim in ws.column_dimensions.items():
            if dim.width:
                analysis.column_widths[key] = float(dim.width)
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                formula = cell.value if isinstance(cell.value, str) and cell.value.startswith("=") else None
                if formula:
                    analysis.formulas[cell.coordinate] = formula
                analysis.non_empty_cells.append(
                    TemplateCellInfo(
                        coordinate=cell.coordinate,
                        value=cell.value,
                        formula=formula,
                        style_id=cell.style_id,
                    )
                )
        return analysis
