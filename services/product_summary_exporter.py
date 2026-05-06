from __future__ import annotations

import json
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from services.normalizer import normalize_name
from services.parser import OrderTextParser


@dataclass(slots=True)
class ProductSummaryExportResult:
    path: Path
    sheet_name: str


class ProductSummaryExporter:
    def export(self, raw_json: str, destination: Path, partner_groups: dict[str, str]) -> ProductSummaryExportResult:
        payload = json.loads(raw_json)
        order = payload.get("orderData")
        if not isinstance(order, dict):
            raise ValueError("В JSON не найден объект orderData.")

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Лист1"
        self._setup_sheet(worksheet)
        self._render_sheet(worksheet, order, partner_groups)
        workbook.save(destination)
        return ProductSummaryExportResult(destination, worksheet.title)

    def _render_sheet(self, ws, order: dict, partner_groups: dict[str, str]) -> None:
        ws["A1"] = f"Заказ номер: {order.get('code', '')}"
        ws["A1"].font = Font(bold=True)

        row = 3
        for cart, is_root in self._iter_carts(order):
            entries = self._extract_product_entries(cart)
            if not entries:
                continue
            ws[f"A{row}"] = self._build_cart_header(cart, partner_groups, is_root=is_root)
            ws[f"D{row}"] = self._partner_comment(cart, partner_groups)
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"D{row}"].font = Font(bold=True)
            row += 1
            for entry in entries:
                ws.cell(row, 1, self._entry_code(entry))
                ws.cell(row, 2, self._entry_name(entry))
                ws.cell(row, 3, self._entry_quantity(entry))
                row += 1

    def _setup_sheet(self, ws) -> None:
        ws.column_dimensions["A"].width = 8.78
        ws.column_dimensions["B"].width = 32.78
        ws.column_dimensions["C"].width = 5.55
        ws.column_dimensions["D"].width = 51.78

    def _iter_carts(self, order: dict):
        yield order, True
        for subcart in order.get("subCarts") or []:
            if isinstance(subcart, dict):
                yield subcart, False

    def _extract_product_entries(self, cart: dict) -> list[dict]:
        result: list[dict] = []
        for entry in cart.get("entries") or []:
            if not isinstance(entry, dict):
                continue
            product = entry.get("product") or {}
            if product.get("type") == "SERVICE" or product.get("lynxServiceType") == "REGISTRATION_FEE":
                continue
            result.append(entry)
        return result

    def _build_cart_header(self, cart: dict, partner_groups: dict[str, str], *, is_root: bool) -> str:
        ordered_by = cart.get("orderedBy") or {}
        full_name = OrderTextParser._extract_person_name(ordered_by)
        group_name = partner_groups.get(normalize_name(full_name), "")
        account_code = (
            str((ordered_by.get("account") or {}).get("code") or "")
            or str((cart.get("account") or {}).get("code") or "")
        )
        amount = self._money_value(cart.get("subTotal"))
        weight = self._measure_weight(cart)
        points = self._decimal_value(cart.get("pointValue"))

        if is_root:
            header_core = f"{account_code} Ваши Позиции"
        elif full_name:
            header_core = f"{account_code} {full_name}".strip()
        elif group_name:
            header_core = f"Группа: {group_name}"
        else:
            header_core = account_code or "Подкорзина"

        if group_name:
            header_core = f"{header_core} ({group_name})"
        return (
            f"         {header_core} "
            f"Сумма: {self._fmt_number(amount)} Масса: {self._fmt_number(weight)} Баллы: {self._fmt_number(points)}"
        )

    def _partner_comment(self, cart: dict, partner_groups: dict[str, str]) -> str:
        ordered_by = cart.get("orderedBy") or {}
        full_name = OrderTextParser._extract_person_name(ordered_by)
        normalized_name = normalize_name(full_name)
        return partner_groups.get(f"{normalized_name}__comment", "")

    def _entry_code(self, entry: dict) -> str:
        product = entry.get("product") or {}
        code = str(product.get("alias") or product.get("code") or "").strip()
        return code

    def _entry_name(self, entry: dict) -> str:
        product = entry.get("product") or {}
        return str(product.get("name") or "").strip()

    def _entry_quantity(self, entry: dict):
        quantity = entry.get("quantity")
        if isinstance(quantity, float) and quantity.is_integer():
            return int(quantity)
        return quantity

    def _measure_weight(self, cart: dict) -> Decimal:
        cis_measure = cart.get("cisMeasure") or {}
        grand_measure = cart.get("grandTotalMeasure") or {}
        return self._decimal_value(cis_measure.get("weight")) or self._decimal_value(grand_measure.get("weight"))

    @staticmethod
    def _money_value(value) -> Decimal:
        if isinstance(value, dict):
            value = value.get("value")
        return ProductSummaryExporter._decimal_value(value)

    @staticmethod
    def _decimal_value(value) -> Decimal:
        try:
            if value in (None, ""):
                return Decimal("0")
            return Decimal(str(value))
        except (InvalidOperation, TypeError, ValueError):
            return Decimal("0")

    @staticmethod
    def _fmt_number(value: Decimal) -> str:
        normalized = value.normalize()
        text = format(normalized, "f")
        if "." in text:
            text = text.rstrip("0").rstrip(".")
        if text == "-0":
            text = "0"
        if "." not in text and value == value.to_integral_value():
            return f"{value:.1f}"
        return text
