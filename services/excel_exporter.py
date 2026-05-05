from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from models import Order, OrderItem


HEADERS = [
    "№",
    "ФИО",
    "Сумма заказа в тенге",
    "Рег. взнос",
    "Скидка тенге",
    "Сумма с учетом скидки",
    "Сумма в рублях",
    "Доставка",
    "Сумма с доставкой",
    "Сумма в рублях факт",
    "Доставка факт",
    "Сумма с доставкой факт",
    "Пришло в тенге",
    "Пришло в рублях",
    "Возврат по товару",
    "Перевели",
    "Дата заказа",
    "Курс тенге",
    "Курс тенге факт",
    "Расходы",
]

COLUMN_WIDTHS = {
    "A": 3.2,
    "B": 22.7,
    "C": 10.7,
    "D": 6.2,
    "E": 9.5,
    "F": 9.0,
    "G": 9.2,
    "H": 8.7,
    "I": 10.0,
    "J": 9.2,
    "K": 8.7,
    "L": 10.2,
    "M": 8.2,
    "N": 8.2,
    "O": 9.2,
    "P": 9.2,
    "Q": 11.0,
    "R": 5.2,
    "S": 5.7,
    "T": 8.5,
    "Y": 1.2,
}

BLUE = "FF2E75B5"
RED = "FFC00000"
GREEN = "FF00B050"
NUM_FORMAT = "#,##0"
TEXT_ALIGN = Alignment(horizontal="left")
CENTER_ALIGN = Alignment(horizontal="center")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

HEADER_COLOR_BY_COL = {
    "C": BLUE,
    "F": BLUE,
    "J": BLUE,
    "M": BLUE,
    "P": BLUE,
    "D": RED,
    "G": RED,
    "I": RED,
    "L": RED,
    "N": RED,
    "O": RED,
    "E": GREEN,
    "H": GREEN,
    "K": GREEN,
    "T": GREEN,
}

BODY_COLOR_BY_COL = {
    3: BLUE,
    6: BLUE,
    10: BLUE,
    13: BLUE,
    16: BLUE,
    4: RED,
    7: RED,
    9: RED,
    12: RED,
    14: RED,
    15: RED,
    5: GREEN,
    8: GREEN,
    11: GREEN,
    20: GREEN,
}


@dataclass(slots=True)
class ExportResult:
    path: Path
    sheet_name: str


class ExcelExporter:
    def export_new(self, order: Order, destination: Path) -> ExportResult:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = self._make_sheet_name(order, workbook.sheetnames)
        self._render_sheet(worksheet, order)
        workbook.save(destination)
        return ExportResult(destination, worksheet.title)

    def append_sheet(self, order: Order, source_file: Path, destination: Path | None = None) -> ExportResult:
        workbook = load_workbook(source_file)
        sheet_name = self._make_sheet_name(order, workbook.sheetnames)
        worksheet = workbook.create_sheet(title=sheet_name)
        self._render_sheet(worksheet, order)
        target = destination or source_file
        workbook.save(target)
        return ExportResult(target, sheet_name)

    def _render_sheet(self, ws: Worksheet, order: Order) -> None:
        self._setup_sheet(ws)
        self._write_headers(ws)
        self._write_parameters(ws, order)

        row = 3
        total_rows: list[int] = []
        for group_name, items in self._group_items(order):
            ws.cell(row, 2, group_name).font = Font(bold=True)
            row += 1
            start_row = row
            for index, item in enumerate(items, start=1):
                self._write_item_row(ws, row, index, item)
                row += 1
            self._write_group_total(ws, row, start_row, row - 1)
            total_rows.append(row)
            row += 1

        grand_total_row = row + 1
        self._write_grand_total(ws, grand_total_row, total_rows)
        self._finalize_parameter_formulas(ws, grand_total_row)
        self._write_footer(ws, order, grand_total_row + 1, total_rows)

    def _setup_sheet(self, ws: Worksheet) -> None:
        ws.freeze_panes = "A2"
        for column, width in COLUMN_WIDTHS.items():
            ws.column_dimensions[column].width = width
        for hidden_column in ("Y",):
            ws.column_dimensions[hidden_column].hidden = True

    def _write_headers(self, ws: Worksheet) -> None:
        for index, title in enumerate(HEADERS, start=1):
            cell = ws.cell(1, index, title)
            color = HEADER_COLOR_BY_COL.get(get_column_letter(index))
            cell.font = Font(bold=True, color=color) if color else Font(bold=True)
            cell.alignment = HEADER_ALIGN
        ws.row_dimensions[1].height = 42

    def _write_parameters(self, ws: Worksheet, order: Order) -> None:
        ws["H2"] = float(self._to_fraction(order.delivery_percent))
        ws["Q2"] = order.order_date.strftime("%d.%m.%Y") if order.order_date else ""
        ws["R2"] = float(order.tenge_rate)
        ws["S2"] = float(order.tenge_rate_fact)
        ws["T3"] = float(order.expenses)

    def _write_item_row(self, ws: Worksheet, row: int, index: int, item: OrderItem) -> None:
        ws[f"A{row}"] = index
        ws[f"B{row}"] = item.full_name
        ws[f"C{row}"] = float(item.amount_tenge)
        if item.registration_fee is not None:
            ws[f"D{row}"] = float(item.registration_fee)
        ws[f"E{row}"] = float(item.discount_tenge)
        ws[f"M{row}"] = float(item.amount_tenge)
        if item.received_rub is not None:
            ws[f"N{row}"] = float(item.received_rub)
        if item.transferred_rub is not None:
            ws[f"Y{row}"] = float(item.transferred_rub)
        self._apply_item_formulas(ws, row)
        self._style_body_row(ws, row)

    def _apply_item_formulas(self, ws: Worksheet, row: int) -> None:
        ws[f"F{row}"] = f"=C{row}-E{row}"
        ws[f"G{row}"] = f"=F{row}/$R$2"
        ws[f"H{row}"] = f"=(C{row}-D{row})/$R$2*$H$2"
        ws[f"I{row}"] = f"=G{row}+H{row}"
        ws[f"J{row}"] = f"=F{row}/$S$2"
        ws[f"K{row}"] = f"=G{row}*$K$2"
        ws[f"L{row}"] = f"=J{row}+K{row}"
        if ws[f"N{row}"].value is None:
            ws[f"N{row}"] = f"=(C{row}*$H$2+F{row})/$R$2"
        ws[f"O{row}"] = f"=(C{row}-M{row})/$R$2"

    def _write_group_total(self, ws: Worksheet, row: int, start_row: int, end_row: int) -> None:
        ws[f"B{row}"] = "Всего:"
        for column in "CDEFGHIJKLMNO":
            ws[f"{column}{row}"] = f"=SUM({column}{start_row}:{column}{end_row})"
        ws[f"P{row}"] = f"=SUM(Y{start_row}:Y{end_row})"
        self._style_total_row(ws, row)

    def _write_grand_total(self, ws: Worksheet, row: int, total_rows: list[int]) -> None:
        ws[f"B{row}"] = "Итого за весь заказ:"
        for column in "CDEFGHIJKLMNO":
            refs = ",".join(f"{column}{group_row}" for group_row in total_rows)
            ws[f"{column}{row}"] = f"=SUM({refs})" if refs else 0
        transferred_refs = ",".join(f"P{group_row}" for group_row in total_rows)
        ws[f"P{row}"] = f"=SUM({transferred_refs})" if transferred_refs else 0
        self._style_total_row(ws, row)

    def _finalize_parameter_formulas(self, ws: Worksheet, grand_total_row: int) -> None:
        ws["K2"] = f"=(G{grand_total_row}+T2)/G{grand_total_row}-1"
        ws["T2"] = f"=SUM(T3:T{grand_total_row-1})"

    def _write_footer(self, ws: Worksheet, order: Order, start_row: int, total_rows: list[int]) -> None:
        ws[f"B{start_row}"] = "Номер заказа:"
        ws[f"C{start_row}"] = order.order_number
        ws[f"B{start_row + 1}"] = "Город отправки:"
        ws[f"C{start_row + 1}"] = order.dispatch_city
        ws[f"B{start_row + 2}"] = "Отправитель:"
        ws[f"C{start_row + 2}"] = order.sender

        label_row = start_row + 3
        paid_row = start_row + 4
        balance_row = start_row + 5
        fact_balance_row = start_row + 6
        bcc_row = start_row + 7
        bcc_balance_row = start_row + 8
        grand_total_row = start_row - 1

        footer_labels = {
            "C": "Всего тенге",
            "E": "Мы оплатили",
            "F": "Рубли курс",
            "G": "Рубли курс факт",
            "H": "Разница",
            "I": "Доставка %",
            "L": "Расходы на доставку",
            "N": "1% Ед",
            "O": "Доставка Едрышовых",
        }
        for column, value in footer_labels.items():
            ws[f"{column}{label_row}"] = value
            ws[f"{column}{label_row}"].font = Font(bold=True)

        second_group_total = total_rows[1] if len(total_rows) > 1 else None
        first_group_total = total_rows[0] if total_rows else None

        ws[f"B{paid_row}"] = "Оплатил:"
        ws[f"C{paid_row}"] = f"=F{grand_total_row}"
        ws[f"E{paid_row}"] = f"=C{paid_row}"
        ws[f"F{paid_row}"] = f"=E{paid_row}/R2"
        ws[f"G{paid_row}"] = f"=E{paid_row}/S2"
        ws[f"H{paid_row}"] = f"=F{paid_row}-G{paid_row}"
        ws[f"I{paid_row}"] = f"=H{grand_total_row}"
        ws[f"L{paid_row}"] = f"=K{grand_total_row}"
        if first_group_total and second_group_total:
            ws[f"N{paid_row}"] = f"=(H{first_group_total}+H{second_group_total})/6"
            ws[f"O{paid_row}"] = f"=H{second_group_total}-N{paid_row}"
        else:
            ws[f"N{paid_row}"] = 0
            ws[f"O{paid_row}"] = 0

        ws[f"B{balance_row}"] = "Остаток:"
        ws[f"C{balance_row}"] = f"=H{paid_row}+I{paid_row}-L{paid_row}-N{paid_row}"
        ws[f"O{balance_row}"] = f"=I{second_group_total}-N{paid_row}" if second_group_total else 0

        ws[f"B{fact_balance_row}"] = "Факт. Остаток:"
        ws[f"C{fact_balance_row}"] = f"=E{bcc_row}/S2+P{grand_total_row}-G{paid_row}"

        ws[f"B{bcc_row}"] = "Оплатил через BCC:"
        ws[f"E{bcc_row}"] = 0

        ws[f"B{bcc_balance_row}"] = "Остаток:"
        ws[f"E{bcc_balance_row}"] = f"=E{paid_row}-E{bcc_row}"
        ws[f"G{bcc_balance_row}"] = f"=E{bcc_balance_row}/S2"

        for row in range(start_row, bcc_balance_row + 1):
            ws[f"B{row}"].font = Font(bold=True)
        self._style_footer_values(ws, paid_row, balance_row, fact_balance_row, bcc_row, bcc_balance_row)

    def _style_body_row(self, ws: Worksheet, row: int) -> None:
        for col in range(1, 21):
            cell = ws.cell(row, col)
            cell.alignment = TEXT_ALIGN if col == 2 else CENTER_ALIGN
            color = BODY_COLOR_BY_COL.get(col)
            if color:
                cell.font = Font(color=color)
            if col >= 3:
                cell.number_format = NUM_FORMAT

    def _style_total_row(self, ws: Worksheet, row: int) -> None:
        for col in range(1, 21):
            cell = ws.cell(row, col)
            color = BODY_COLOR_BY_COL.get(col)
            cell.font = Font(bold=True, color=color) if color else Font(bold=True)
            cell.alignment = TEXT_ALIGN if col == 2 else CENTER_ALIGN
            if col >= 3:
                cell.number_format = NUM_FORMAT

    def _style_footer_values(
        self,
        ws: Worksheet,
        paid_row: int,
        balance_row: int,
        fact_balance_row: int,
        bcc_row: int,
        bcc_balance_row: int,
    ) -> None:
        footer_colors = {
            "C": BLUE,
            "E": GREEN,
            "F": BLUE,
            "G": BLUE,
            "H": RED,
            "I": GREEN,
            "L": GREEN,
            "N": RED,
            "O": RED,
        }
        rows = [paid_row, balance_row, fact_balance_row, bcc_row, bcc_balance_row]
        for row in rows:
            for column, color in footer_colors.items():
                cell = ws[f"{column}{row}"]
                if cell.value is None:
                    continue
                cell.font = Font(color=color)
                cell.alignment = CENTER_ALIGN
                cell.number_format = NUM_FORMAT

    def _group_items(self, order: Order) -> list[tuple[str, list[OrderItem]]]:
        grouped: dict[str, list[OrderItem]] = {}
        for item in order.items:
            group_name = item.group_name or "Без группы"
            grouped.setdefault(group_name, []).append(item)
        return [(name, grouped[name]) for name in sorted(grouped)]

    @staticmethod
    def _to_fraction(value) -> float:
        numeric = float(value)
        return numeric / 100 if numeric > 1 else numeric

    @staticmethod
    def _make_sheet_name(order: Order, existing_names: list[str]) -> str:
        base = order.order_date.strftime("%Y.%m.%d") if order.order_date else (order.order_number or "Заказ")
        if order.order_number:
            base = f"{base}_{order.order_number}"
        candidate = base[:31]
        if candidate not in existing_names:
            return candidate
        index = 2
        while True:
            suffix = f"({index})"
            alt = f"{base[:31-len(suffix)]}{suffix}"
            if alt not in existing_names:
                return alt
            index += 1
